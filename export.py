"""
Export job results to a formatted .xlsx file.

Columns: Applied? | Title | Company | Location | Date Posted | Source | Link
- "Applied?" pre-filled with "No" — change to "Yes" as you apply
- "Link" column contains clickable hyperlinks
- Header row is bold and frozen
- Column widths are auto-fitted
"""

from pathlib import Path
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


OUTPUT_DIR = Path("output")

# Header background colours per region
REGION_COLORS = {
    "sweden": "006AA7",   # Swedish blue
    "eu":     "003399",   # EU blue
    "all":    "2C3E50",   # dark slate
}


def save_xlsx(df: pd.DataFrame, region: str = "all") -> Path:
    OUTPUT_DIR.mkdir(exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    path = OUTPUT_DIR / f"jobs_{region}_{timestamp}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Jobs"

    header_color = REGION_COLORS.get(region, "2C3E50")
    header_font  = Font(bold=True, color="FFFFFF")
    header_fill  = PatternFill("solid", fgColor=header_color)

    # ── Headers ────────────────────────────────────────────────────────────────
    columns = ["Applied?", "Title", "Company", "Location", "Date Posted", "Source", "Link"]
    for col_idx, header in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font   = header_font
        cell.fill   = header_fill
        cell.alignment = Alignment(horizontal="center")

    ws.freeze_panes = "A2"   # freeze header row

    # ── Rows ───────────────────────────────────────────────────────────────────
    df_cols = {
        "title":       "Title",
        "company":     "Company",
        "location":    "Location",
        "date_posted": "Date Posted",
        "site":        "Source",
        "job_url":     "Link",
    }

    for row_idx, (_, row) in enumerate(df.iterrows(), 2):
        # Applied? column
        ws.cell(row=row_idx, column=1, value="No")

        ws.cell(row=row_idx, column=2, value=str(row.get("title",       "")))
        ws.cell(row=row_idx, column=3, value=str(row.get("company",     "")))
        ws.cell(row=row_idx, column=4, value=str(row.get("location",    "")))
        ws.cell(row=row_idx, column=5, value=str(row.get("date_posted", "")))
        ws.cell(row=row_idx, column=6, value=str(row.get("site",        "")))

        # Clickable hyperlink in Link column
        url = str(row.get("job_url", ""))
        link_cell = ws.cell(row=row_idx, column=7, value="Open" if url else "")
        if url:
            link_cell.hyperlink = url
            link_cell.font      = Font(color="0563C1", underline="single")

    # ── Column widths ──────────────────────────────────────────────────────────
    widths = [10, 45, 30, 25, 18, 12, 10]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    wb.save(path)
    return path
